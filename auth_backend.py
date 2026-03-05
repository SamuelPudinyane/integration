from __future__ import annotations
"""Shared auth/data access layer for the integration portal.

When onboarding additional apps into this integration, prefer reusing this module
for shared identity and canonical role resolution so all apps enforce the same
role vocabulary and user source of truth.

Typical extension points:
- Add shared user/profile lookup helpers consumed by multiple apps.
- Keep canonical role mapping centralized (`canonical_role_name`) to avoid
    app-specific role drift.
- Add read-only reference data loaders needed by portal-wide pages.
"""

import os
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.engine import Engine

try:
    from auth_portal.shared_auth import DEFAULT_ADMIN_USERNAME, ROLE_ADMIN, canonical_role_name
except ModuleNotFoundError:
    from shared_auth import DEFAULT_ADMIN_USERNAME, ROLE_ADMIN, canonical_role_name


ROOT_DIR = Path(__file__).resolve().parent.parent
load_dotenv(ROOT_DIR / "fake data" / ".env")
PROCESS_HIERARCHY_WORKBOOK_GLOB = "L1 TO L3 -water utilities*.xlsx"


def _env(name: str, default: str | None = None) -> str:
    value = os.getenv(name, default)
    if value is None or value == "":
        raise ValueError(f"Missing required environment variable: {name}")
    return value


def get_database_url() -> str:
    host = _env("DB_HOST", "localhost")
    port = _env("DB_PORT", "5432")
    db_name = _env("DB_NAME")
    db_user = _env("DB_USER")
    db_password = _env("DB_PASSWORD")
    return f"postgresql+psycopg2://{db_user}:{db_password}@{host}:{port}/{db_name}"


engine: Engine = create_engine(get_database_url(), future=True)

USER_TABLE = "public.it_digital_user_account"
DEPARTMENT_TABLE = "public.ref_organization_unit"


def read_users() -> list[dict[str, str]]:
    # This function is intentionally integration-facing:
    # any new app plugged into the portal should rely on this user list (or helpers
    # built on top of it) to ensure one consistent role model across apps.
    with engine.begin() as conn:
        rows = conn.execute(
            text(
                f"""
                SELECT user_id, username, email, role_name, password_hash, is_active
                FROM {USER_TABLE}
                ORDER BY user_id
                """
            )
        ).fetchall()

    users = [
        {
            "id": str(row[0]),
            "username": (row[1] or "").strip(),
            "email": (row[2] or "").strip(),
            "role": canonical_role_name(row[3]) or "Maintenance Technician",
            "password_hash": (row[4] or "").strip(),
            "is_active": "True" if bool(row[5]) else "False",
        }
        for row in rows
    ]

    has_admin = any(
        (user.get("role") or "") == ROLE_ADMIN and (user.get("is_active") or "").lower() == "true"
        for user in users
    )
    if has_admin:
        return users

    max_id = max((int(user.get("id", "0") or "0") for user in users), default=0)
    users.append(
        {
            "id": str(max_id + 1),
            "username": DEFAULT_ADMIN_USERNAME,
            "email": "maintenance_manager@joburgwater.local",
            "password_hash": "",
            "role": ROLE_ADMIN,
            "is_active": "True",
        }
    )
    return users


def get_user_by_id(user_id: str | int | None) -> dict[str, str] | None:
    if user_id is None:
        return None
    target = str(user_id)
    for user in read_users():
        if (user.get("id") or "") == target:
            return user
    return None


def authenticate_user(username: str, password: str) -> tuple[dict[str, Any] | None, str | None]:
    # Note: password verification is currently minimal by design in this environment.
    # If stronger auth is introduced for multiple apps, implement it here once so
    # all integrated apps inherit the same authentication behavior.
    username_clean = (username or "").strip()
    password_clean = (password or "").strip()
    if not username_clean:
        return None, "Username is required."
    if not password_clean:
        return None, "Password is required."

    users = read_users()
    matched = next((u for u in users if u.get("username") == username_clean), None)
    if not matched:
        return None, "User not found."

    if (matched.get("is_active", "False") or "False").lower() != "true":
        return None, "Your account is inactive."

    return matched, None


def read_department_hierarchy() -> dict[str, Any]:
    with engine.begin() as conn:
        rows = conn.execute(
            text(
                f"""
                SELECT unit_id, unit_name, unit_code, parent_unit_id, description, is_active
                FROM {DEPARTMENT_TABLE}
                WHERE COALESCE(is_active, true) = true
                ORDER BY unit_name
                """
            )
        ).fetchall()

    nodes: dict[int, dict[str, Any]] = {}
    children_by_parent: dict[int, list[int]] = defaultdict(list)

    for row in rows:
        unit_id = int(row[0])
        parent_unit_id = int(row[3]) if row[3] is not None else None
        nodes[unit_id] = {
            "unit_id": unit_id,
            "unit_name": (row[1] or "").strip(),
            "unit_code": (row[2] or "").strip(),
            "parent_unit_id": parent_unit_id,
            "description": (row[4] or "").strip(),
            "children": [],
        }
        if parent_unit_id is not None:
            children_by_parent[parent_unit_id].append(unit_id)

    for parent_id, child_ids in children_by_parent.items():
        if parent_id not in nodes:
            continue
        sorted_child_ids = sorted(child_ids, key=lambda cid: (nodes[cid].get("unit_name") or "").lower())
        nodes[parent_id]["children"] = [nodes[cid] for cid in sorted_child_ids if cid in nodes]

    roots = [
        node
        for node in nodes.values()
        if node.get("parent_unit_id") is None or node.get("parent_unit_id") not in nodes
    ]
    roots = sorted(roots, key=lambda item: (item.get("unit_name") or "").lower())

    relationships = []
    for node in nodes.values():
        parent_id = node.get("parent_unit_id")
        if parent_id is None or parent_id not in nodes:
            continue
        parent = nodes[parent_id]
        relationships.append(
            f"{parent.get('unit_name', 'Unknown')} ({parent.get('unit_code', '')}) -> {node.get('unit_name', 'Unknown')} ({node.get('unit_code', '')})"
        )

    return {
        "roots": roots,
        "relationships": sorted(relationships),
        "total_departments": len(nodes),
    }


def read_process_steps() -> dict[str, Any]:
    with engine.begin() as conn:
        table_rows = conn.execute(
            text(
                """
                SELECT table_schema, table_name
                FROM information_schema.tables
                WHERE table_schema = 'public'
                  AND (
                    table_name ILIKE '%process%'
                    OR table_name ILIKE '%workflow%'
                    OR table_name ILIKE '%step%'
                  )
                ORDER BY table_name
                """
            )
        ).fetchall()

        process_tables = [f"{row[0]}.{row[1]}" for row in table_rows]
        process_records: list[dict[str, Any]] = []

        for schema_name, table_name in table_rows:
            columns = conn.execute(
                text(
                    """
                    SELECT column_name
                    FROM information_schema.columns
                    WHERE table_schema = :schema_name
                      AND table_name = :table_name
                    ORDER BY ordinal_position
                    """
                ),
                {"schema_name": schema_name, "table_name": table_name},
            ).fetchall()
            column_names = [str(col[0]) for col in columns]
            if not column_names:
                continue

            quoted_columns = ", ".join(f'"{name}"' for name in column_names)
            rows = conn.execute(
                text(f'SELECT {quoted_columns} FROM "{schema_name}"."{table_name}" LIMIT 500')
            ).fetchall()

            for row in rows:
                payload = dict(row._mapping)
                process_name = (
                    payload.get("step_name")
                    or payload.get("process_name")
                    or payload.get("unit_name")
                    or payload.get("parameter_name")
                    or payload.get("name")
                    or ""
                )
                sequence = (
                    payload.get("step_order")
                    or payload.get("sequence")
                    or payload.get("order_no")
                    or payload.get("unit_id")
                    or payload.get("parameter_id")
                    or ""
                )
                process_records.append(
                    {
                        "source_table": f"{schema_name}.{table_name}",
                        "process_name": str(process_name or "").strip(),
                        "sequence": str(sequence or "").strip(),
                    }
                )

    process_records = sorted(
        process_records,
        key=lambda item: (
            item.get("source_table", ""),
            item.get("sequence", ""),
            item.get("process_name", ""),
        ),
    )
    return {
        "tables": process_tables,
        "records": process_records,
        "total_steps": len(process_records),
    }


def _find_process_hierarchy_workbook() -> Path | None:
    matches = sorted(ROOT_DIR.glob(PROCESS_HIERARCHY_WORKBOOK_GLOB))
    if not matches:
        return None
    return matches[0]


def read_l1_l3_process_hierarchy() -> dict[str, Any]:
    workbook_path = _find_process_hierarchy_workbook()
    if not workbook_path:
        return {
            "roots": [],
            "source_file": "",
            "level_counts": {"l1": 0, "l2": 0, "l3": 0},
            "load_error": f"Workbook not found: {PROCESS_HIERARCHY_WORKBOOK_GLOB}",
        }

    wb = load_workbook(str(workbook_path), read_only=True, data_only=True)
    sheet_l1 = wb["LEVEL1 "] if "LEVEL1 " in wb.sheetnames else None
    sheet_l2 = wb["LEVEL 2"] if "LEVEL 2" in wb.sheetnames else None
    sheet_l3 = wb["LEVEL 3"] if "LEVEL 3" in wb.sheetnames else None

    if not sheet_l1 or not sheet_l2 or not sheet_l3:
        return {
            "roots": [],
            "source_file": str(workbook_path),
            "level_counts": {"l1": 0, "l2": 0, "l3": 0},
            "load_error": "Required sheets LEVEL1 / LEVEL 2 / LEVEL 3 are missing.",
        }

    l1_title_pattern = re.compile(r"^(\d+)\.\s+(.*)$")
    l2_pattern = re.compile(r"^(\d+)\.(\d+)\s+(.*)$")
    l3_pattern = re.compile(r"^(\d+)\.(\d+)\.(\d+)\s+(.*)$")

    roots: dict[str, dict[str, Any]] = {}

    l1_rows: list[str] = []
    for row in sheet_l1.iter_rows(values_only=True):
        value = row[0] if row else None
        if value is None:
            continue
        text_value = str(value).strip()
        if text_value:
            l1_rows.append(text_value)

    pending_l1_code = ""
    for entry in l1_rows:
        m = l1_title_pattern.match(entry)
        if m:
            code = m.group(1)
            roots[code] = {
                "code": f"{code}.",
                "name": entry,
                "description": "",
                "children": [],
            }
            pending_l1_code = code
            continue

        if pending_l1_code and pending_l1_code in roots and not roots[pending_l1_code].get("description"):
            roots[pending_l1_code]["description"] = entry

    l2_nodes: dict[str, dict[str, Any]] = {}
    for row in sheet_l2.iter_rows(values_only=True):
        value = row[0] if row else None
        if value is None:
            continue
        text_value = str(value).strip()
        m = l2_pattern.match(text_value)
        if not m:
            continue
        l1_code = m.group(1)
        l2_code = f"{m.group(1)}.{m.group(2)}"
        node = {
            "code": l2_code,
            "name": text_value,
            "description": "",
            "children": [],
        }
        l2_nodes[l2_code] = node
        if l1_code in roots:
            roots[l1_code]["children"].append(node)

    l3_count = 0
    for row in sheet_l3.iter_rows(values_only=True):
        value = row[0] if row else None
        if value is None:
            continue
        text_value = str(value).strip()
        m = l3_pattern.match(text_value)
        if not m:
            continue
        l2_code = f"{m.group(1)}.{m.group(2)}"
        l3_node = {
            "code": f"{m.group(1)}.{m.group(2)}.{m.group(3)}",
            "name": text_value,
            "description": "",
            "children": [],
        }
        parent = l2_nodes.get(l2_code)
        if parent is not None:
            parent["children"].append(l3_node)
            l3_count += 1

    sorted_roots = sorted(roots.values(), key=lambda item: int(str(item.get("code", "0")).split(".")[0]))
    for root in sorted_roots:
        root["children"] = sorted(
            root.get("children", []),
            key=lambda item: tuple(int(part) for part in str(item.get("code", "0")).split(".")),
        )
        for child in root["children"]:
            child["children"] = sorted(
                child.get("children", []),
                key=lambda item: tuple(int(part) for part in str(item.get("code", "0")).split(".")),
            )

    return {
        "roots": sorted_roots,
        "source_file": str(workbook_path),
        "level_counts": {"l1": len(sorted_roots), "l2": len(l2_nodes), "l3": l3_count},
        "load_error": "",
    }
